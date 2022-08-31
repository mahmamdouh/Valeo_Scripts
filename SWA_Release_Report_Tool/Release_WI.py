import connectors.polarion_connector as c
from openpyxl import Workbook
from openpyxl.styles import Font
import xlsxwriter
import os
import re
from datetime import datetime
#from GUI import Task
from openpyxl.chart import BarChart, Reference, Series

Release_WI_Warning_Message ={
    "Message" : ""
}

def get_input_data():
    info_list = []

    # Get data from Input_Data_File.txt
    with open("Input_Data_File.txt") as InfoFile:
        for line in InfoFile:
            info_list.append(line[(line.find(':') + 1):].rstrip('\n'))

    i = 0
    for element in info_list:
        info_list[i] = element.replace(" ", "")
        i += 1
    #print(info_list)
    return info_list


def create_output_directory():
    output_path = os.getcwd() + "\Outputs"

    # Check existence of Output folder and create it if not.
    if not os.path.exists(output_path):
        os.makedirs(output_path)
        #print(output_path + ' : created')

    return output_path


def get_folder_data(info_list):
    #print("Connecting to Polarion...")
    not_connected = True

    # Connect to Polarion database
    while not_connected:
        try:
            # Connect using the username and password
            polarion_object = c.Polarion("https://vseapolarion.vnet.valeo.com/polarion")
            polarion_object.connect(str(info_list[0]), str(info_list[1]))

            folder_content = ''

            folder_content = polarion_object.tracker_webservice.service.getModules(info_list[2], info_list[3])
            not_connected = False
        except BaseException as e:
            # #print connection error message
            #print(str(e))
            #print("Retry...")
            pass
        except Exception as e:
            # #print connection error message
            #print(str(e))
            #print("Retry...")
            pass
        except:
            #print("Retry...")
            pass

    ##print("Folder content downloaded")

    # Disconnect Polarion.
    polarion_object.disconnect()


    return folder_content


def get_work_items_ids(folder_content,info_list):
    #print("Step : 0")
    docs_title_list = []
    docs_content = {}
    #print("Step : 1")
    # Get documents title in list and make dictionary for IDs list for each document.
    for element in folder_content:
        docs_title_list.append(element.title)
        docs_content[element.title] = element.homePageContent

    # Loop on each document and get it's IDs, then save it in the dictionary.
    #print("Step : 2")
    for element in docs_title_list:
        home_page_content = docs_content[element].content

        res = []
        ids_list = []

        # Search for IDs in the page data.
        for elements in re.finditer("params=id=", home_page_content):
            # #print(home_page_content[elements.end():elements.end() + 16])
            if info_list[2] == "optimus":
                res.append(home_page_content[elements.end():elements.end() + 11])
            elif info_list[2] == "model_kit":
                res.append(home_page_content[elements.end():elements.end() + 15])
            elif info_list[2] == "mma_cm1e_dcdc":
                res.append(home_page_content[elements.end():elements.end() + 19])
            elif info_list[2] == "VW_MEB_Inverter":
                res.append(home_page_content[elements.end():elements.end() + 17])
            elif info_list[2] == "VW_MEB_Inverter_Base_Minus":
                res.append(home_page_content[elements.end():elements.end() + 23])
            elif info_list[2] == "PMA1":
                res.append(home_page_content[elements.end():elements.end() + 11])
            elif info_list[2] == "me_s230":
                res.append(home_page_content[elements.end():elements.end() + 11])

        # Filter duplicated IDs.
        for i in range(len(res)):
            # Check if the string ends with digits (number) and return None if not.
            for j in range(3):
                last_character = re.search(r'\d+$', res[i])
                if last_character is None:
                    res[i] = res[i][:-1]
                j += 1
            if res[i] not in ids_list:
                ids_list.append(res[i])

        # Save the IDs list in the dictionary.
        docs_content[element] = ids_list
    #print("Step : 3")
    return docs_content


def get_work_items_data(docs_content,info_list):
    docs_content_detail = {}
    #Release_WI["Status"] = "Connecting to Polarion..."
    #print("Connecting to Polarion...")

    # Connect using the username and password
    polarion_object = c.Polarion("https://vseapolarion.vnet.valeo.com/polarion")
    polarion_object.connect(str(info_list[0]), str(info_list[1]))

    # Get workitem details for each document.
    for element in docs_content:
        ids_list = docs_content[element]
        not_connected = True

        # Connect to Polarion database
        while not_connected:
            try:
                # Clear the list data to avoid any corrupted data.
                work_item_list = []
                ids_group = '('

                # Get WorkItems data.
                for i in range(len(ids_list)):
                    ids_group += (ids_list[i] + ' ')
                ids_group = ids_group[:-1]
                ids_group += ')'

                query = str("project.id:" + info_list[2] + " AND id:" + ids_group)

                work_item_list.append(polarion_object.tracker_webservice.service.queryWorkItems(
                    query, "priority", ["id", "title", "type", "status", "customFields.safety.KEY",
                                        "customFields.variant.KEY", "customFields.architecture_type.KEY"]))
                not_connected = False
            except ConnectionError as e:
                # #print connection error message
                #print(str(e))
                #print("Retry...")
                pass
            except Exception as e:
                # #print connection error message
                #print(str(e))
                #print("Retry...")
                pass
            except BaseException as e:
                # #print connection error message
                #print(str(e))
                #print("Retry...")
                pass
            except:
                #print("Retry...")
                pass

        docs_content_detail[element] = work_item_list
        #Task["Name"] = "Document" + str(element)+ "' data downloaded successfully"
        #print("Document '" + element + "' data downloaded successfully")

    ##print("Documents workitems data downloaded")

    # Disconnect Polarion.
    polarion_object.disconnect()

    return docs_content_detail



def RWI_data_write(docs_content_detail, folder_content,info_list,output_path):
    File_Name = "\Traceability_Matrix_Data_" + info_list[3] + ".xlsx"
    output_workbook = output_path + File_Name
    workbook1 = xlsxwriter.Workbook(output_workbook)

    # Create Excel document

    ws_Exe = workbook1.add_worksheet('Execution Details')
    ws_kpi  = workbook1.add_worksheet("KPI")
    ws_interface = workbook1.add_worksheet('SW Interfaces')
    ws_hsi = workbook1.add_worksheet('HSI Elements')
    ws_static = workbook1.add_worksheet('Static Views')
    ws_dynamic = workbook1.add_worksheet('Dynamic Views')
    ws_runnable = workbook1.add_worksheet('Runnables')
    ws_SWC = workbook1.add_worksheet('swComponent')
    ws_DWI = workbook1.add_worksheet('diagnostic')
    ws_REQ = workbook1.add_worksheet('softwareRequirement')
    ws_others = workbook1.add_worksheet('Generic')


    workbook1 ,ws_kpi, ws_interface, ws_hsi, ws_static, ws_dynamic, ws_runnable,ws_SWC,ws_DWI,ws_REQ, ws_others,ws_Exe= set_sheets_headers\
        (workbook1,ws_Exe ,ws_kpi, ws_interface, ws_hsi, ws_static, ws_dynamic, ws_runnable,ws_SWC,ws_DWI,ws_REQ, ws_others,info_list)

    kpi = {
        "SW Interfaces": [{"total": 0, "released": 0, "not released": 0}],
        "HSI Elements": [{"total": 0, "released": 0, "not released": 0}],
        "Static Views": [{"total": 0, "released": 0, "not released": 0}],
        "Dynamic Views": [{"total": 0, "released": 0, "not released": 0}],
        "Runnables": [{"total": 0, "released": 0, "not released": 0}],
        "swComponent": [{"total": 0, "released": 0, "not released": 0}],
        "diagnostic": [{"total": 0, "released": 0, "not released": 0}],
        "softwareRequirement": [{"total": 0, "released": 0, "not released": 0}],
        "Generic": [{"total": 0, "released": 0, "not released": 0}],
    }

    # Pages Indix
    Interfaces_Indix = 1
    HSI_Elements_Indix= 1
    Static_Views_Indix= 1
    Dynamic_Views_Indix= 1
    Runnables_Indix= 1
    swComponent_Indix= 1
    diagnostic_Indix= 1
    softwareRequirement_Indix= 1
    Generic_Indix= 1
    WI_Type=""
    index =0
    ws = ws_Exe
    Variant_Selected ="base+/base-"
    #print("Pass here 6")
    # Get Workitems data.
    # check variant selection
    #print("Variant Input:",info_list[5] )
    if info_list[5] == "NOT variant.KEY:base\+":
        Variant_Selected = "Not base+"
    elif info_list[5] == "NOT variant.KEY:base\-":
        Variant_Selected = "Not base-"
        #print("Base - only selected")
    elif info_list[5] == "NOT variant.KEY: (base\+ base\-)":
        Variant_Selected = "Not base+/base-"
    elif info_list[5] == "variant.KEY:base\+":
        Variant_Selected = "base+"
    elif info_list[5] == "variant.KEY:base\-":
        Variant_Selected = "base-"
    elif info_list[5] == "variant.KEY: (base\+ base\-)":
        Variant_Selected = "base+/base-"
    else:
        Variant_Selected = "base+/base-"

    for elements in docs_content_detail:
        Document_Location = str(info_list[3])+"/"+elements
        Document_Title = elements
        for element in docs_content_detail[elements]:
            for sub_element in element:
                id = sub_element.id
                title = sub_element.title
                status = sub_element.status.id
                type = sub_element.type.id


                if type == "hw_sw_md_architecture" or type == "sw_interface" or type == "diagnostic" or type == "softwareRequirement" or type == "swComponent":
                    #print("WI Variant"  , variant_C )
                    ##print("Pass in for loop 6")
                    variant = ''
                    arch_type = ''

                    try:
                        if type == 'hw_sw_md_architecture':
                            for item in sub_element.customFields.Custom:
                                if item.key == "architecture_type":
                                    arch_type = item.value.id

                    except:
                        pass

                    try:
                        for items in sub_element.customFields.Custom:
                            if items.key == "variant":
                                for item in items.value.EnumOptionId:
                                    variant += item.id
                    except:
                        pass
                    ##print("Pass in for loop 7")
                    if type == "sw_interface":
                        ws = ws_interface
                        Interfaces_Indix = Interfaces_Indix +1
                        index = Interfaces_Indix
                        WI_Type = "SW Interfaces"
                        arch_type = "SW Interfaces"


                    elif type == "diagnostic":
                        ##print("DIAG:",id,title)
                        diagnostic_Indix = diagnostic_Indix+1
                        ws =ws_DWI
                        index = diagnostic_Indix
                        WI_Type = "diagnostic"

                    elif type == "softwareRequirement":
                        ##print("REQ:", id, title)
                        softwareRequirement_Indix = softwareRequirement_Indix+1
                        ws =ws_REQ
                        index = softwareRequirement_Indix
                        WI_Type = "softwareRequirement"

                    elif type == "swComponent":
                        ##print("SWC:", id, title)
                        swComponent_Indix = swComponent_Indix+1
                        ws =ws_SWC
                        index = swComponent_Indix
                        WI_Type = "swComponent"
                    elif type == "hw_sw_md_architecture":
                        if arch_type == "hsi_element":
                            HSI_Elements_Indix = HSI_Elements_Indix+1
                            ws = ws_hsi
                            index = HSI_Elements_Indix
                            WI_Type = "HSI Elements"
                        elif arch_type == "hw_static_view" or arch_type == "hw_md_static_view" or \
                                arch_type == "hw_sw_static_view" or arch_type == "hw_sw_md_static_view" or \
                                arch_type == "md_static_view" or arch_type == "sw_static_view":
                            ws = ws_static
                            Static_Views_Indix = Static_Views_Indix+1
                            index = Static_Views_Indix
                            WI_Type = "Static Views"

                        elif arch_type == "sw_dynamic_view":
                            ws = ws_dynamic
                            Dynamic_Views_Indix = Dynamic_Views_Indix+1
                            index = Dynamic_Views_Indix
                            WI_Type = "Dynamic Views"

                        elif arch_type == "sw_runnable":
                            ws = ws_runnable
                            Runnables_Indix = Runnables_Indix+1
                            index = Runnables_Indix
                            WI_Type = "Runnables"

                        else:
                            ws = ws_others
                            Generic_Indix = Generic_Indix+1
                            index = Generic_Indix
                            WI_Type = "Generic"
                            arch_type = "Generic"
                    ##print("Pass in for loop 8")

                    a = 'A' + str(index)
                    b = 'B' + str(index)
                    c = 'C' + str(index)
                    d = 'D' + str(index)
                    e = 'E' + str(index)
                    f = 'F' + str(index)
                    g = 'G' + str(index)
                    h = 'H' + str(index)
                    i = 'I' + str(index)
                    ##print("Pass in for loop 9")
                    # Write the data in the Excel sheet
                    link = link_generate(info_list, id, "workitem")
                    URL = str(link)
                    #ws.cell(row=index, column=1).value =Get_Hyper_Link(link, id)
                    ws.write_url(a, URL, string=id)
                    ws.write(b, str(title))
                    ws.write(c, str(type))
                    ws.write(d, str(status))

                    ws.write(e, str(variant))
                    ##print("Pass in for loop 10")
                    #for folder_element in folder_content:
                       # if folder_element.moduleName == elements:
                           # location = folder_element.moduleLocation
                            #print("Location", location)
                    ##print("Pass in for loop 11")

                    link = link_generate(info_list, Document_Location, "document")
                    URL = str(link)
                    #ws.cell(row=index, column=6).value = Get_Hyper_Link(link, elements)
                    ws.write_url(f, URL, string=str(Document_Title))


                    safety = """=IF(OR(REGEXMATCH(B""" + str(index) + """, "Sfty.*"),REGEXMATCH(B""" + str(index) + """, "ActvDcha.*")), "Yes", "No")"""
                    ws.write(g, str(safety))
                    delta = """=IF(OR(REGEXMATCH(F""" + str(index) + """ , "CtrlEm"), REGEXMATCH(F""" + str(index) + """ , "DetmnEmRotorTemp"), REGEXMATCH(F""" + str(index) + """ , "GenSysSply"), REGEXMATCH(F""" + str(index) + """ , "DetmnSafeTq"), REGEXMATCH(F""" + str(index) + """ , "ActeSafeSt")), "Yes", "No")"""
                    ws.write(h, str(delta))
                    ws.write(i, arch_type)
                    #print("Pass here 6.000")
                    kpi = get_kpi(kpi, WI_Type, status)


                else :
                    pass
        print("Document ",Document_Title , "Finished")

    #print("Pass here 7")
    workbook1 ,ws_kpi= write_kpi_data(kpi, workbook1, ws_kpi,info_list)
    #print("Pass here 8")
    # Save Excel sheet
    while True:
        try:
            workbook1.close()
            #os.startfile(output_workbook)
            break
        except:
            print("Message from Component : ",Release_WI_Warning_Message["Message"] )
            Release_WI_Warning_Message["Message"] = "File Is already opened ! , please close Excel file "
            pass






def set_sheets_headers(workbook1,ws_Exe, ws_kpi, ws_interface, ws_hsi, ws_static, ws_dynamic, ws_runnable,ws_SWC,ws_DWI,ws_REQ, ws_others,info_list):
    ws_list = [ws_interface, ws_hsi, ws_static, ws_dynamic, ws_runnable,ws_SWC,ws_DWI,ws_REQ, ws_others]
    #print("Pass here 1 ")
    now = datetime.now()  # current date and time
    Date_Data = now.strftime("%m/%d/%Y, %H:%M:%S")

    cell_format = workbook1.add_format()
    cell_format.set_bold()

    ws_Exe.write("A1", "Exection Date", cell_format)
    ws_Exe.write("A2", "Generated By", cell_format)
    ws_Exe.write("A3", "SWA Baseline", cell_format)
    ws_Exe.write("A4", "SWREQ Baseline", cell_format)
    #ws_Exe.write("A5", "Variant", cell_format)

    ws_Exe.write("B1", str(Date_Data))
    ws_Exe.write("B2", str(info_list[0]))
    ws_Exe.write("B3", str(info_list[3]))
    ws_Exe.write("B4", str(info_list[4]))
    #ws_Exe.write("B5", str(info_list[5]))

    #print("Pass here 4 ")
    for element in ws_list:
        ws = element
        ws.autofilter('A1:I5000')
        ws.write("A1", "IDs")
        ws.write("B1", "Title")
        ws.write("C1", "Type")
        ws.write("D1", "Status")
        ws.write("E1", "Variant")
        ws.write("F1", "System Function")
        ws.write("G1", "Safety")
        ws.write("H1", "Delta")
        ws.write("I1", "Architecture Type")


    ws = ws_kpi
    ws.write("A2", "Total")
    ws.write("A3", "Released")
    ws.write("A4", "Not Released")

    ws.write("B1", "SW Interfaces")
    ws.write("C1", "HSI Elements")
    ws.write("D1", "Static Views")
    ws.write("E1", "Dynamic Views")
    ws.write("F1", "Runnables")
    ws.write("G1", "SWC")
    ws.write("H1", "DWI")
    ws.write("I1", "Req")
    ws.write("J1", "Generic")



    #print("Pass here 5 ")
    return workbook1,ws_kpi, ws_interface, ws_hsi, ws_static, ws_dynamic, ws_runnable,ws_SWC,ws_DWI,ws_REQ, ws_others,ws_Exe


def link_generate(info_list, id, link_type):
    if link_type == "workitem":
        link = "https://vseapolarion.vnet.valeo.com/polarion/#/project/" + info_list[2] + "/workitem?id=" + id
    elif link_type == "testrun":
        link = "https://vseapolarion.vnet.valeo.com/polarion/#/project/" + info_list[2] + "/testrun?id=" + id
    elif link_type == "document":
        link = "https://vseapolarion.vnet.valeo.com/polarion/#/project/" + info_list[2] + "/wiki/" + id

    return link


def get_kpi(kpi, type, status):
    for element in kpi:
        if element == type:
            kpi[type][0]["total"] +=1
            if status == "released":
                kpi[type][0]["released"] += 1
            else:
                kpi[type][0]["not released"] += 1

    return kpi


def write_kpi_data(kpi, wb, ws_kpi,info_list):
    # Write SW Interfaces data.
    #print("Pass here 7.1")
    #ws_kpi.cell(row=2, column=2).value = kpi["SW Interfaces"][0]["total"]
    ws_kpi.write("B2", Query_HyberLink(info_list, "SW Interfaces", kpi["SW Interfaces"][0]["total"]))
    ws_kpi.write("B3",str(kpi["SW Interfaces"][0]["released"]))
    ws_kpi.write("B4", str(kpi["SW Interfaces"][0]["not released"]))
    #print("Pass here 2.1")

    # Write HSI Elements data.
    ws_kpi.write("C2", str(Query_HyberLink(info_list, "(hsi_element", kpi["HSI Elements"][0]["total"])))
    ws_kpi.write("C3", str(kpi["HSI Elements"][0]["released"]))
    ws_kpi.write("C4", str(kpi["HSI Elements"][0]["not released"]))


    # Write Static Views data.
    ws_kpi.write("D2", str(Query_HyberLink(info_list, "sw_static_view", kpi["Static Views"][0]["total"])))
    ws_kpi.write("D3", str(kpi["Static Views"][0]["released"]))
    ws_kpi.write("D4", str(kpi["Static Views"][0]["not released"]))

    # Write Dynamic Views data.
    ws_kpi.write("E2", str(Query_HyberLink(info_list, "sw_dynamic_view", kpi["Dynamic Views"][0]["total"])))
    ws_kpi.write("E3", str(kpi["Dynamic Views"][0]["released"]))
    ws_kpi.write("E4", str(kpi["Dynamic Views"][0]["not released"]))

    # Write Runnables data.
    ws_kpi.write("F2", str(Query_HyberLink(info_list, "sw_runnable", kpi["Runnables"][0]["total"])))
    ws_kpi.write("F3", str(kpi["Runnables"][0]["released"]))
    ws_kpi.write("F4", str(kpi["Runnables"][0]["not released"]))


    # Write Runnables data.
    ws_kpi.write("G2", str(Query_HyberLink(info_list, "swComponent", kpi["swComponent"][0]["total"])))
    ws_kpi.write("G3",  str(kpi["swComponent"][0]["released"]))
    ws_kpi.write("G4",  str(kpi["swComponent"][0]["not released"]))

    # Write Runnables data.
    ws_kpi.write("H2", str(Query_HyberLink(info_list, "diagnostic", kpi["diagnostic"][0]["total"])))
    ws_kpi.write("H3", str(kpi["diagnostic"][0]["released"]))
    ws_kpi.write("H4", str(kpi["diagnostic"][0]["not released"]))

    # Write Runnables data.
    ws_kpi.write("I2", str(Query_HyberLink(info_list, "softwareRequirement", kpi["softwareRequirement"][0]["total"])))
    ws_kpi.write("I3", str(kpi["softwareRequirement"][0]["released"]))
    ws_kpi.write("I4", str(kpi["softwareRequirement"][0]["not released"]))

    # write Others data.
    ws_kpi.write("J2", str(kpi["Generic"][0]["total"]))
    ws_kpi.write("J2", str(kpi["Generic"][0]["released"]))
    ws_kpi.write("J2", str(kpi["Generic"][0]["not released"]))

    #print("Pass here 3.1 finish")
    return wb ,ws_kpi

def Query_HyberLink(info_list , Type , Num):

    '''
    if Type == "SW Interfaces":
        Link = "https://vseapolarion.vnet.valeo.com/polarion/#/project/VW_MEB_Inverter/workitems?query=type%3A"+ Type+ \
        "%20AND%20SQL%3A(select%20WI.C_PK%20from%20MODULE%20M%20inner%20join%20REL_MODULE_WORKITEM%20RMW%20ON%20RMW.FK_URI_MODULE%20%3D%20M.C_URI%20inner%20join%20WORKITEM%20WI%20on%20WI.C_URI%20%3D%20RMW.FK_URI_WORKITEM%20where%20\
        (M.C_LOCATION%20like%20'%25"+info_list[4] + "%25'%20))%20AND%20variant.KEY%3Abase%5C%2B"
    elif Type == "diagnostic":
        Link = "https://vseapolarion.vnet.valeo.com/polarion/#/project/VW_MEB_Inverter/workitems?query=type%3A" + Type + \
               "%20AND%20SQL%3A(select%20WI.C_PK%20from%20MODULE%20M%20inner%20join%20REL_MODULE_WORKITEM%20RMW%20ON%20RMW.FK_URI_MODULE%20%3D%20M.C_URI%20inner%20join%20WORKITEM%20WI%20on%20WI.C_URI%20%3D%20RMW.FK_URI_WORKITEM%20where%20\
               (M.C_LOCATION%20like%20'%25" + info_list[4] + "%25'%20))%20AND%20variant.KEY%3Abase%5C%2B"
    elif Type == "softwareRequirement":
        Link = "https://vseapolarion.vnet.valeo.com/polarion/#/project/VW_MEB_Inverter/workitems?query=type%3A" + Type + \
               "%20AND%20SQL%3A(select%20WI.C_PK%20from%20MODULE%20M%20inner%20join%20REL_MODULE_WORKITEM%20RMW%20ON%20RMW.FK_URI_MODULE%20%3D%20M.C_URI%20inner%20join%20WORKITEM%20WI%20on%20WI.C_URI%20%3D%20RMW.FK_URI_WORKITEM%20where%20\
               (M.C_LOCATION%20like%20'%25" + info_list[4] + "%25'%20))%20AND%20variant.KEY%3Abase%5C%2B"
    elif Type == "swComponent":
        Link = "https://vseapolarion.vnet.valeo.com/polarion/#/project/VW_MEB_Inverter/workitems?query=type%3A" + Type + \
               "%20AND%20SQL%3A(select%20WI.C_PK%20from%20MODULE%20M%20inner%20join%20REL_MODULE_WORKITEM%20RMW%20ON%20RMW.FK_URI_MODULE%20%3D%20M.C_URI%20inner%20join%20WORKITEM%20WI%20on%20WI.C_URI%20%3D%20RMW.FK_URI_WORKITEM%20where%20\
               (M.C_LOCATION%20like%20'%25" + info_list[4] + "%25'%20))%20AND%20variant.KEY%3Abase%5C%2B"

    else:
        Link = "https://vseapolarion.vnet.valeo.com/polarion/#/project/VW_MEB_Inverter/workitems?query=type%3Ahw_sw_md_architecture"\
               "%20AND%20SQL%3A(select%20WI.C_PK%20from%20MODULE%20M%20inner%20join%20REL_MODULE_WORKITEM%20RMW%20ON%20RMW.FK_URI_MODULE%20%3D%20M.C_URI%20inner%20join%20WORKITEM%20WI%20on%20WI.C_URI%20%3D%20RMW.FK_URI_WORKITEM%20where%20\
               (M.C_LOCATION%20like%20'%25" + info_list[4] + "%25'%20))%20AND%20variant.KEY%3Abase%5C%2B%20AND%20architecture_type.KEY%3A"+Type
    #print("Pass here *10")
    HyperLink = "=HYPERLINK(\"" + Link + "\",\"" + str(Num) + "\")"
    #print("HyperLink : ",HyperLink)
    '''
    return Num



'''
if __name__ == '__main__':

    info_list = get_input_data()
    # Create Script Output folder.

    output_path = create_output_directory()

    # Connect to Polarion.
    folder_content = get_folder_data(info_list)
    print("Done 3")
    # For each document get workitems IDs and data.
    docs_content = get_work_items_ids(folder_content,info_list)
    print("Done 4")
    docs_content_detail = get_work_items_data(docs_content,info_list)
    print("Done 5")
    ##print(docs_content_detail)
    # Create Excel sheet and write data.

    RWI_data_write(docs_content_detail, folder_content,info_list,output_path)
    print("Done 6")

'''
