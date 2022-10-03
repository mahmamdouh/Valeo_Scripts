from shutil import copyfile
import xml.etree.ElementTree as ET
import re
import xlwt 
from xlwt import Workbook
import os
from os import listdir
import xlsxwriter
import xlrd
import subprocess
from xlwt import Workbook
from openpyxl import load_workbook
from Get_Stored_Data_Pandas import Get_Release_WIs
from PolarionPlanAPIs import Get_Gerrit_WI_Data
from RTE_API import Log_File


# this component is responsible for parse log file and get all data from it
# and generate diff logs for changes and get the actula functional changes

# parse log file and get its data
def writeInToExcel(wb,ws,sheet,n):
    r = 2
    current_directory = os.getcwd()
    with open(current_directory+'\SubFiles\log.txt') as f:
        #print("Log txt opened ")
        lines = f.readlines()
        for i in range(0,len(lines)):
            commitflag = 0        
            if(lines[i].startswith('commit')):
                commitArr = lines[i].split(' ')
                commitID = commitArr[1]
                codechangesArr = []
                ws.cell(r, column=1).value = commitID

                # find Rolling Build tag
                if (re.findall("tag:", str(lines[i]))):
                    result = re.search('tag:(.*)\)', str(lines[i]))
                    commitTag = result.group(0)
                    ws.cell(r, column=9).value = commitTag
                else:
                    ws.cell(r, column=9).value = "-"
                i = i + 1
            if(lines[i].startswith('Author')):
                result = re.search(':(.*)<', lines[i])
                author = result.group(1)           
                ws.cell(r, column=2).value = author
                i = i + 1
            if (str(lines[i]).startswith("Date")):
                result = re.search(':(.*)+', lines[i])
                commitDate = result.group(0)
                commitDate = commitDate.split('+')
                Date = commitDate[0]
                #print(Date)
                Date = Date.split(':   ')
                ws.cell(r, column=8).value = str(Date[1])


            if (lines[i] == '\n' or str(lines[i]).startswith(' ')):
                i = i + 1
                
            if('  $' in str(lines[i])):
                start = lines[i].index('$')
                end = lines[i].index('$',start+1)

                substring = lines[i][start+1:end]
                ws.cell(r, column=4).value = substring
                # parse Commit message
                Commit_Message = lines[i].split('$')
                ws.cell(r, column=10).value = Commit_Message [2]

            if ('Revert "$' in str(lines[i])):
                start = lines[i].index('$')
                end = lines[i].index('$', start + 1)

                substring = lines[i][start + 1:end]
                ws.cell(r, column=4).value = substring
                # parse Commit message
                Commit_Message = lines[i].split('$')
                ws.cell(r, column=10).value = Commit_Message[2]
                
            if('Change-Id' in str(lines[i])):
                File_List = []
                changeArr = lines[i].split(': ')
                changeID = changeArr[1]
                ws.cell(r, column=3).value = changeID
                i = i + 2
                removedup = []
                while(1):
                    if(i<len(lines)):
                        if(not(str(lines[i]).startswith(' ') or str(lines[i]).startswith('\n'))):
                            #File_Name = lines[i].strip()
                            File_List.append(str(lines[i]))
                            SWCsArr = lines[i].split('/')
                            thefile = SWCsArr[(len(SWCsArr)-1)]
                            thefile = thefile.split('.')
                            if (len(thefile) > 1):
                                extention = thefile[1]
                                #print(thefile[0] + '  ,  ' + thefile[1])
                            if (len(SWCsArr)>=3):
                                if(SWCsArr[2].startswith('Components') and (extention == 'c\n' or extention == 'h\n')):
                                    SWCs = SWCsArr[(len(SWCsArr)-3)]
                                    removedup.append(SWCs)
                            i = i + 1
                            if(i == len(lines)):
                                break
                        else:
                            break
                    else:
                        break
                #print('____________________________________________')
                #print("File List:" ,File_List )
                File_Type_List = define_File_List_Types(File_List)
                file_List_String = List_To_String_Multi_Lines(File_List)
                ws.cell(r, column=11).value = file_List_String
                ws.cell(r, column=12).value = File_Type_List[0]
                ws.cell(r, column=13).value = File_Type_List[1]
                ws.cell(r, column=14).value = File_Type_List[2]
                ws.cell(r, column=15).value = File_Type_List[3]
                ws.cell(r, column=16).value = File_Type_List[4]
                removed = list(dict.fromkeys(removedup))
                for k in range(0,len(removed)):
                    if ( k == 0):
                        ws.cell(r, column=5).value = str(removed[k])
                    else:
                        ws.cell(r, column=5).value = str(ws.cell(r, column=5).value) + ',' + str(removed[k])
                        
                r = r + 1
    return wb ,ws ,sheet
#-------


# this function conver list of strings to one string multi lines
def List_To_String_Multi_Lines(List):
    STR_List =""
    for elements in List:
        STR_List = STR_List + " " + str(elements)

    return STR_List

# this function parse file List and defeine if change has par files , source files , header files , RTRT files
def define_File_List_Types(File_List):
    Rtn_Array =[]
    Rtn_Array =['-','-','-','-','-']
    for elements in File_List:
        if (re.findall("\.par\n", elements)):
            result = re.search('\.par\n', elements)
            Rtn_Array[0] = "Yes"
        if (re.findall("\.a2l\n", elements)):
            result = re.search('\.a2l]n', elements)
            Rtn_Array[1] = "Yes"
        if (re.findall("\.c\n", elements)):
            result = re.search('\.c\n', elements)
            Rtn_Array[2] = "Yes"
        if (re.findall("\.h\n", elements)):
            result = re.search('\.h\n', elements)
            Rtn_Array[3] = "Yes"

        Line_Split = elements.split('/')
        if Line_Split[0] == "test":
            Rtn_Array[4] = "Yes"
    return Rtn_Array


# this function is to create diff logs
def splitCommits(sheet,n,w,mypath,List_Of_Tags ,Rows , SWC_List):

    while  n < Rows+1:
        if SWC_List[n-1] != 0:
            #print("SWC List :",SWC_List[n-1],":")
            try:
                pr = subprocess.Popen(['git', 'show', List_Of_Tags[n-1], ] , cwd = mypath , shell = True, stdout = subprocess.PIPE, stderr = subprocess.PIPE )
                (out, error) = pr.communicate()
                outarray = str(out).split('\\n')
                #print(len(outarray))
                with open('SubFiles/commitLog'+str(w)+'.txt','w') as f:
                    for j in range(0,len(outarray)):
                        f.write(outarray[j])
                        f.write('\n')
                    w = w + 1
            except:
                pass
            #print(w)
        n = n + 1
#--------

# this function is to parse diff logs
def getAffcSWCs(ws, sheet, n ,Rows , SWC_List):
    r = 2
    w = 1
    flag = 0
    bad_exception =0
    #print("Parsing",n ,sheet.nrows)
    while n < Rows+1:

        #print("Parsing afected SWCs function 2")
        #print("Commit log number:", w,SWC_List[n-1])
        if not(SWC_List[n-1] == 0):
            #print("Commit log number:",w,"===================================================================")
            #print("Parsing afected SWCs function 1")
            try :
                with open('SubFiles/commitLog' + str(w) + '.txt') as f:
                    #print("Parsing afected SWCs function")
                    affectedSWCsArr = []
                    flag = 0
                    lines = f.readlines()
                    for i in range(0, len(lines) - 1):
                        Line_Code = str(lines[i])

                        flag = 0
                        if (lines[i].startswith('diff --git')):
                            while (1):
                                i = i + 1
                                #print(i, lines[i])
                                try :
                                    if (lines[i].startswith('+++') or lines[i].startswith('---')):
                                      #print(i, lines[i])
                                        break
                                except:
                                    bad_exception =1
                                    break


                            while (1):
                                if bad_exception == 1:
                                    bad_exception =0
                                    break
                                if (i == len(lines)):
                                    break
                                if (lines[i].startswith('+++') or lines[i].startswith('---')):
                                    NSWCsArr = lines[i].split('/')
                                    thefile = NSWCsArr[(len(NSWCsArr) - 1)]
                                    thefile = thefile.split('.')
                                    if (len(thefile) > 1):
                                        extention = thefile[1]
                                    if (len(NSWCsArr) > 3):
                                        if (NSWCsArr[3].startswith('Components') and (
                                                extention.startswith('c') or extention.startswith('h'))):
                                            CheckSWC = NSWCsArr[(len(NSWCsArr) - 3)]
                                        else:
                                            break
                                if (lines[i].startswith('diff')):
                                    break
                                # if(lines[i].startswith('+/*') or lines[i].startswith('-/*')):
                                # print('--------------->>> Comment <<<---------------')
                                mystring = lines[i]
                                # if(lines[i] == '+ ' or lines[i] == '- '):
                                # print('--------------->>> EMPTY LINE <<<---------------')
                                if ((lines[i].startswith('+') and len(lines[i]) > 2)):
                                    if ((lines[i].startswith('+') and mystring[1] != '/' and (
                                            mystring[2] != '*' or mystring[2] != '/')) or (
                                            lines[i].startswith('-') and mystring[1] != '/' and (
                                            mystring[2] != '*' or mystring[2] != '/'))):
                                        if ((lines[i].startswith('+') and mystring[1] != '+' and mystring[2] != '+') or (
                                                lines[i].startswith('-') and mystring[1] != '-' and mystring[2] != '-')):
                                            flag = 1
                                            # print("CHANGEEE ----->>> " + lines[i])
                                if (flag == 1):
                                    affectedSWCsArr.append(CheckSWC)
                                i = i + 1
            except:
                pass
            removed = list(dict.fromkeys(affectedSWCsArr))
            #print("removed list:" + str(removed))
            # if (len(removed) != 0):
            for k in range(0, len(removed)):
                if (k == 0):
                    ws.cell(r, column=6).value = str(removed[k])
                else:
                    ws.cell(r, column=6).value = str(ws.cell(r, column=6).value) + ',' + str(removed[k])

            r = r + 1
            # else:
            # ws.cell(r, column=6).value =  "No SWC Affected"
            # r = r + 1

            # r = r + 1
            w = w + 1

        else:
            r = r + 1
            # n = n + 1
        n = n + 1
        # r = r + 1

# generate compatable report
def filter_Report(Report_Dict):
    List_Of_Tags = []
    List_Of_splitted_Commits = []
    SWC_List =[]
    Tags_List =[]
    #print("Pandas:" , Report_Dict)
    Rows = len(Report_Dict["SWCs"])
    for elements in Report_Dict["SWCs"]:
        SWC_List.append(Report_Dict["SWCs"][elements])
        if Report_Dict["SWCs"][elements]== 0:
            pass
        else:
            List_Of_splitted_Commits.append(Report_Dict["SWCs"][elements])
    for elements in Report_Dict["SWCs"]:
        SWC_List.append(Report_Dict["SWCs"][elements])
        if Report_Dict["SWCs"][elements]== "nan":
            pass
        else:
            List_Of_Tags.append(Report_Dict["commit"][elements])

    for elements in Report_Dict["Tag"]:
        if Report_Dict["Tag"][elements]== "nan":
            pass
        else:
            Tags_List.append(Report_Dict["Tag"][elements])


    return List_Of_Tags,Rows , SWC_List ,len(List_Of_splitted_Commits),Tags_List

# main Runnable
def Gerrit_Changes_Analysis_Runnable(Input_Data):

    mypath = Input_Data["L_Repo"]
    #print("Repo Path",mypath)
    wb = load_workbook('SubFiles\Integration_Report.xlsx')
    ws = wb.active
    read_file = xlrd.open_workbook('SubFiles\Integration_Report.xlsx')
    sheet = read_file.sheet_by_index(0)


    n = 1
    affectedSWCsArr = []
    SWC = ""
    flag = 0
    NSWCsArr = []
    w = 1
    #print("Write to Excel file ")
    writeInToExcel(wb,ws,sheet,n)
    wb.save('SubFiles\Integration_Report.xlsx')
    #print("Excel file saved ")
    current_directory = os.getcwd()
    Report_Dict = Get_Release_WIs(current_directory + "\SubFiles\Integration_Report.xlsx")
    List_Of_Tags ,Rows , SWC_List , Splitted_Commits , Tags_List= filter_Report(Report_Dict)
    Log_File["ID"] = Tags_List
    #print("commites to splited ")
    splitCommits(sheet,n,w,mypath,List_Of_Tags ,Rows , SWC_List)
    #print("commites splited ")
    getAffcSWCs(ws, sheet, n , Rows , SWC_List)
    #print("getAffcSWCs generated")
    wb.save('SubFiles\Integration_Report.xlsx')
    #print("Report saved")

    return Tags_List


