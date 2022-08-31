from shutil import copyfile
import xml.etree.ElementTree as ET
import re
import xlwt 
from xlwt import Workbook
import os
from os import listdir
import xlsxwriter
import xlrd
import subprocess
from xlwt import Workbook
from openpyxl import load_workbook
from Get_Stored_Data_Pandas import Get_Release_WIs



def writeInToExcel(wb,ws,sheet,n):
    r = 2
    current_directory = os.getcwd()
    with open(current_directory+'\SubFiles\log.txt') as f:
        #print("Log txt opened ")
        lines = f.readlines()
        for i in range(0,len(lines)):
            commitflag = 0        
            if(lines[i].startswith('commit')):
                commitArr = lines[i].split(' ')
                commitID = commitArr[1]
                codechangesArr = []
                ws.cell(r, column=1).value = commitID
                i = i + 1
            if(lines[i].startswith('Author')):
                result = re.search(':(.*)<', lines[i])
                author = result.group(1)           
                ws.cell(r, column=2).value = author
                i = i + 1
                
            if(str(lines[i]).startswith("Date") or lines[i] == '\n' or str(lines[i]).startswith(' ')):
                i = i + 1
                
            if('  $' in str(lines[i])):
                start = lines[i].index('$')
                end = lines[i].index('$',start+1)

                substring = lines[i][start+1:end]
                ws.cell(r, column=4).value = substring
                
            if('Change-Id' in str(lines[i])):
                changeArr = lines[i].split(': ')
                changeID = changeArr[1]
                ws.cell(r, column=3).value = changeID
                i = i + 2
                removedup = []
                while(1):
                    if(i<len(lines)):
                        if(not(str(lines[i]).startswith(' ') or str(lines[i]).startswith('\n'))):
                            SWCsArr = lines[i].split('/')
                            thefile = SWCsArr[(len(SWCsArr)-1)]
                            thefile = thefile.split('.')
                            if (len(thefile) > 1):
                                extention = thefile[1]
                                #print(thefile[0] + '  ,  ' + thefile[1])
                            if (len(SWCsArr)>=3):
                                if(SWCsArr[2].startswith('Components') and (extention == 'c\n' or extention == 'h\n')):
                                    SWCs = SWCsArr[(len(SWCsArr)-3)]
                                    removedup.append(SWCs)
                            i = i + 1
                            if(i == len(lines)):
                                break
                        else:
                            break
                    else:
                        break
                #print('____________________________________________')    
                removed = list(dict.fromkeys(removedup))
                for k in range(0,len(removed)):
                    if ( k == 0):
                        ws.cell(r, column=5).value = str(removed[k])
                    else:
                        ws.cell(r, column=5).value = str(ws.cell(r, column=5).value) + ',' + str(removed[k])
                        
                r = r + 1
    return wb ,ws ,sheet
#-------
def splitCommits(sheet,n,w,mypath,List_Of_Tags ,Rows , SWC_List):

    while  n < Rows+1:
        if SWC_List[n-1] != 0:
            #print("SWC List :",SWC_List[n-1],":")
            pr = subprocess.Popen(['git', 'show', List_Of_Tags[n-1], ] , cwd = mypath , shell = True, stdout = subprocess.PIPE, stderr = subprocess.PIPE )
            (out, error) = pr.communicate()
            outarray = str(out).split('\\n')
            #print(len(outarray))
            with open('SubFiles/commitLog'+str(w)+'.txt','w') as f:
                for j in range(0,len(outarray)):
                    f.write(outarray[j])
                    f.write('\n')
                w = w + 1

            #print(w)
        n = n + 1
#--------

def getAffcSWCs(ws, sheet, n ,Rows , SWC_List):
    r = 2
    w = 1
    flag = 0
    bad_exception =0
    #print("Parsing",n ,sheet.nrows)
    while n < Rows+1:

        #print("Parsing afected SWCs function 2")
        #print("Commit log number:", w,SWC_List[n-1])
        if not(SWC_List[n-1] == 0):
            #print("Commit log number:",w,"===================================================================")
            #print("Parsing afected SWCs function 1")
            with open('SubFiles/commitLog' + str(w) + '.txt') as f:
                #print("Parsing afected SWCs function")
                affectedSWCsArr = []
                flag = 0
                lines = f.readlines()
                for i in range(0, len(lines) - 1):
                    Line_Code = str(lines[i])

                    flag = 0
                    if (lines[i].startswith('diff --git')):
                        while (1):
                            i = i + 1
                            #print(i, lines[i])
                            try :
                                if (lines[i].startswith('+++') or lines[i].startswith('---')):
                                  #print(i, lines[i])
                                    break
                            except:
                                bad_exception =1
                                break


                        while (1):
                            if bad_exception == 1:
                                bad_exception =0
                                break
                            if (i == len(lines)):
                                break
                            if (lines[i].startswith('+++') or lines[i].startswith('---')):
                                NSWCsArr = lines[i].split('/')
                                thefile = NSWCsArr[(len(NSWCsArr) - 1)]
                                thefile = thefile.split('.')
                                if (len(thefile) > 1):
                                    extention = thefile[1]
                                if (len(NSWCsArr) > 3):
                                    if (NSWCsArr[3].startswith('Components') and (
                                            extention.startswith('c') or extention.startswith('h'))):
                                        CheckSWC = NSWCsArr[(len(NSWCsArr) - 3)]
                                    else:
                                        break
                            if (lines[i].startswith('diff')):
                                break
                            # if(lines[i].startswith('+/*') or lines[i].startswith('-/*')):
                            # print('--------------->>> Comment <<<---------------')
                            mystring = lines[i]
                            # if(lines[i] == '+ ' or lines[i] == '- '):
                            # print('--------------->>> EMPTY LINE <<<---------------')
                            if ((lines[i].startswith('+') and len(lines[i]) > 2)):
                                if ((lines[i].startswith('+') and mystring[1] != '/' and (
                                        mystring[2] != '*' or mystring[2] != '/')) or (
                                        lines[i].startswith('-') and mystring[1] != '/' and (
                                        mystring[2] != '*' or mystring[2] != '/'))):
                                    if ((lines[i].startswith('+') and mystring[1] != '+' and mystring[2] != '+') or (
                                            lines[i].startswith('-') and mystring[1] != '-' and mystring[2] != '-')):
                                        flag = 1
                                        # print("CHANGEEE ----->>> " + lines[i])
                            if (flag == 1):
                                affectedSWCsArr.append(CheckSWC)
                            i = i + 1

            removed = list(dict.fromkeys(affectedSWCsArr))
            #print("removed list:" + str(removed))
            # if (len(removed) != 0):
            for k in range(0, len(removed)):
                if (k == 0):
                    ws.cell(r, column=6).value = str(removed[k])
                else:
                    ws.cell(r, column=6).value = str(ws.cell(r, column=6).value) + ',' + str(removed[k])

            r = r + 1
            # else:
            # ws.cell(r, column=6).value =  "No SWC Affected"
            # r = r + 1

            # r = r + 1
            w = w + 1

        else:
            r = r + 1
            # n = n + 1
        n = n + 1
        # r = r + 1

def filter_Report(Report_Dict):
    List_Of_Tags = []
    List_Of_splitted_Commits = []
    SWC_List =[]
    Rows = len(Report_Dict["SWCs"])
    for elements in Report_Dict["SWCs"]:
        SWC_List.append(Report_Dict["SWCs"][elements])
        if Report_Dict["SWCs"][elements]== 0:
            pass
        else:
            List_Of_splitted_Commits.append(Report_Dict["SWCs"][elements])
    for elements in Report_Dict["SWCs"]:
        SWC_List.append(Report_Dict["SWCs"][elements])
        if Report_Dict["SWCs"][elements]== "nan":
            pass
        else:
            List_Of_Tags.append(Report_Dict["commit"][elements])

    return List_Of_Tags,Rows , SWC_List ,len(List_Of_splitted_Commits)
def Runnable_1(Input_Data):

    mypath = Input_Data["L_Repo"]
    #print("Repo Path",mypath)
    wb = load_workbook('SubFiles\Integration_Report.xlsx')
    ws = wb.active
    read_file = xlrd.open_workbook('SubFiles\Integration_Report.xlsx')
    sheet = read_file.sheet_by_index(0)


    n = 1
    affectedSWCsArr = []
    SWC = ""
    flag = 0
    NSWCsArr = []
    w = 1
    #print("Write to Excel file ")
    writeInToExcel(wb,ws,sheet,n)
    wb.save('SubFiles\Integration_Report.xlsx')
    #print("Excel file saved ")
    current_directory = os.getcwd()
    Report_Dict = Get_Release_WIs(current_directory + "\SubFiles\Integration_Report.xlsx")
    List_Of_Tags ,Rows , SWC_List , Splitted_Commits= filter_Report(Report_Dict)
    #print("commites to splited ")
    splitCommits(sheet,n,w,mypath,List_Of_Tags ,Rows , SWC_List)
    #print("commites splited ")
    getAffcSWCs(ws, sheet, n , Rows , SWC_List)
    #print("getAffcSWCs generated")
    wb.save('SubFiles\Integration_Report.xlsx')
    #print("Report saved")



